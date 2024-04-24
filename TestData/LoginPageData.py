import openpyxl


class DataLoginPage:

    @staticmethod
    def getTestData(testCases):
        book = openpyxl.load_workbook("C:\\OrangeHRM\\DataFiles\\test_data.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCases:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
